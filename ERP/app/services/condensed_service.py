from __future__ import annotations

from calendar import monthrange
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import String, cast, extract, or_
from sqlalchemy.orm import Session, joinedload

from app.models.client import Client
from app.models.product import Product
from app.models.production_order import ProductionOrder
from app.models.quotation import Quotation
from app.models.quotation_item import QuotationItem
from app.services.production_order_service import (
    PRODUCTION_ORDER_STATUSES as PO_STATUS_TUPLES,
    normalize_status,
    transition_status,
)
from app.services.quotation_design_service import get_design_urls, sync_legacy_design_file

CONDENSED_QUOTATION_STATUSES = ["aprobada", "produccion", "en_produccion"]

TRACKING_STATUSES = [(code, label) for code, label in PO_STATUS_TUPLES if code != "cancelado"]

TRACKING_STATUS_LABELS = dict(TRACKING_STATUSES)

IN_PRODUCTION_STATUSES = {"diseno", "produccion", "envio"}

TRACKING_STATUS_COLORS = {
    "pendiente": ("bg-gray-100", "text-gray-700"),
    "diseno": ("bg-blue-100", "text-blue-700"),
    "produccion": ("bg-purple-100", "text-purple-700"),
    "envio": ("bg-cyan-100", "text-cyan-800"),
    "entregado": ("bg-emerald-100", "text-emerald-800"),
}


def _item_is_custom(item: QuotationItem) -> bool:
    if item.product_id is None:
        return True
    product = item.product
    return bool(product and product.custom)


def _order_tracking_status(quotation: Quotation) -> str:
    order = quotation.production_order
    if order and order.status:
        return normalize_status(order.status)
    return "pendiente"


def _order_tracking_meta(quotation: Quotation) -> dict[str, str]:
    order = quotation.production_order
    if order:
        return {
            "assigned_to": order.designer or "",
            "notes": order.notes or order.observations or "",
        }
    return {"assigned_to": "", "notes": ""}


def _order_display_ref(quotation: Quotation | None) -> tuple[int | None, str]:
    """Etiqueta OP real según la orden de producción vinculada."""
    if not quotation:
        return None, "—"
    order = quotation.production_order
    if order and order.id:
        return order.id, f"OP-{order.id:04d}"
    return quotation.id, f"Cotización #{quotation.id}"


def _product_name(item: QuotationItem) -> str:
    if item.detail:
        return item.detail
    if item.product and item.product.name:
        return item.product.name
    return "—"


def _apply_filters(
    query,
    db: Session,
    *,
    month: int | None,
    year: int | None,
    client_id: int | None,
    search: str,
    date_basis: str = "delivery",
):
    query = query.filter(Quotation.status.in_(CONDENSED_QUOTATION_STATUSES))

    if month or year:
        target_year = year or date.today().year
        if date_basis == "quotation":
            query = query.filter(extract("year", Quotation.created_at) == target_year)
            if month:
                query = query.filter(extract("month", Quotation.created_at) == month)
        else:
            if month:
                start = date(target_year, month, 1)
                end = date(target_year, month, monthrange(target_year, month)[1])
            else:
                start = date(target_year, 1, 1)
                end = date(target_year, 12, 31)
            query = query.filter(Quotation.delivery_date.isnot(None))
            query = query.filter(Quotation.delivery_date >= start)
            query = query.filter(Quotation.delivery_date <= end)

    if client_id:
        query = query.filter(Quotation.client_id == client_id)

    if search:
        term = f"%{search.strip()}%"
        matching_quotation_ids = (
            db.query(QuotationItem.quotation_id)
            .join(Quotation, QuotationItem.quotation_id == Quotation.id)
            .join(Client, Quotation.client_id == Client.id)
            .outerjoin(Product, QuotationItem.product_id == Product.id)
            .filter(
                Quotation.status.in_(CONDENSED_QUOTATION_STATUSES),
                or_(
                    Client.name.ilike(term),
                    QuotationItem.detail.ilike(term),
                    QuotationItem.theme.ilike(term),
                    QuotationItem.color.ilike(term),
                    QuotationItem.measure.ilike(term),
                    Product.name.ilike(term),
                    cast(Quotation.id, String).ilike(term),
                ),
            )
            .distinct()
        )
        query = query.filter(Quotation.id.in_(matching_quotation_ids))

    return query


def _order_has_custom(group: dict[str, Any]) -> bool:
    return any(product.get("is_custom") for product in group.get("products", []))


def _order_matches_search(group: dict[str, Any], search: str) -> bool:
    if not search:
        return True

    term = search.strip().lower()
    if not term:
        return True

    haystack = [
        group.get("client_name") or "",
        str(group.get("quotation_id") or ""),
        group.get("order_label") or "",
    ]
    for product in group.get("products", []):
        haystack.extend([
            product.get("product_name") or "",
            product.get("measure") or "",
            product.get("theme") or "",
            product.get("color") or "",
        ])

    return any(term in value.lower() for value in haystack if value)


def filter_order_groups(
    groups: list[dict[str, Any]],
    *,
    production_status: str = "",
    custom_filter: str = "",
    search: str = "",
) -> list[dict[str, Any]]:
    status_filter = (production_status or "").strip().lower()
    custom = (custom_filter or "").strip().lower()

    filtered: list[dict[str, Any]] = []
    for group in groups:
        group_status = (group.get("production_status") or "pendiente").lower()

        if status_filter and group_status != status_filter:
            continue

        has_custom = _order_has_custom(group)
        if custom == "yes" and not has_custom:
            continue
        if custom == "no" and has_custom:
            continue

        if search and not _order_matches_search(group, search):
            continue

        filtered.append(group)

    return filtered


def sort_order_groups(
    groups: list[dict[str, Any]],
    group_by: str = "order",
) -> list[dict[str, Any]]:
    if group_by == "client":
        return sorted(
            groups,
            key=lambda group: (
                (group.get("client_name") or "").lower(),
                -(group.get("delivery_date") or date.min).toordinal()
                if group.get("delivery_date")
                else 0,
                -(group.get("quotation_id") or 0),
            ),
        )
    if group_by == "product":
        return sorted(
            groups,
            key=lambda group: (
                (group["products"][0]["product_name"] if group.get("products") else "").lower(),
                -(group.get("quotation_id") or 0),
            ),
        )
    if group_by == "delivery":
        return sorted(
            groups,
            key=lambda group: (
                group.get("delivery_date") or date.min,
                group.get("quotation_id") or 0,
            ),
            reverse=True,
        )
    return sorted(
        groups,
        key=lambda group: (
            group.get("quotation_date") or datetime.min,
            group.get("quotation_id") or 0,
        ),
        reverse=True,
    )


def annotate_group_sections(
    groups: list[dict[str, Any]],
    group_by: str = "order",
) -> list[dict[str, Any]]:
    annotated: list[dict[str, Any]] = []
    for group in groups:
        section_label = None
        if group_by == "client":
            section_label = group.get("client_name") or "Sin cliente"
        elif group_by == "delivery":
            delivery = group.get("delivery_date")
            section_label = (
                delivery.strftime("%d/%m/%Y") if delivery else "Sin fecha de entrega"
            )
        elif group_by == "product":
            products = group.get("products") or []
            section_label = products[0]["product_name"] if products else "Sin producto"

        annotated.append({**group, "section_label": section_label})

    return annotated


def get_filtered_order_groups(db: Session, **filters) -> list[dict[str, Any]]:
    items = build_condensed_query(db, **filters).all()
    groups = build_order_groups(items)
    groups = filter_order_groups(
        groups,
        production_status=filters.get("production_status", ""),
        custom_filter=filters.get("custom_filter", ""),
        search=filters.get("search", ""),
    )
    group_by = filters.get("group_by", "order")
    groups = sort_order_groups(groups, group_by)
    return annotate_group_sections(groups, group_by)


def _apply_group_order(query, group_by: str):
    if group_by == "client":
        return query.order_by(
            Client.name.asc(),
            Quotation.delivery_date.desc(),
            Quotation.id.desc(),
            QuotationItem.id.desc(),
        )
    if group_by == "product":
        return query.order_by(
            QuotationItem.detail.asc(),
            Quotation.id.desc(),
            QuotationItem.id.desc(),
        )
    return query.order_by(
        Quotation.delivery_date.desc(),
        Quotation.id.desc(),
        QuotationItem.id.desc(),
    )


def build_condensed_query(
    db: Session,
    *,
    month: int | None = None,
    year: int | None = None,
    client_id: int | None = None,
    production_status: str = "",
    custom_filter: str = "",
    search: str = "",
    group_by: str = "order",
    date_basis: str = "delivery",
):
    query = (
        db.query(QuotationItem)
        .join(Quotation, QuotationItem.quotation_id == Quotation.id)
        .join(Client, Quotation.client_id == Client.id)
        .outerjoin(Product, QuotationItem.product_id == Product.id)
        .options(
            joinedload(QuotationItem.quotation).joinedload(Quotation.client),
            joinedload(QuotationItem.quotation).joinedload(Quotation.designs),
            joinedload(QuotationItem.quotation).joinedload(Quotation.production_tracking),
            joinedload(QuotationItem.quotation).joinedload(Quotation.production_order),
            joinedload(QuotationItem.product),
        )
    )

    query = _apply_filters(
        query,
        db,
        month=month,
        year=year,
        client_id=client_id,
        search=search,
        date_basis=date_basis,
    )

    return _apply_group_order(query, group_by)


def build_condensed_row(
    item: QuotationItem,
    *,
    order_number: int | None,
    order_label: str,
    item_number: int = 1,
) -> dict[str, Any]:
    quotation = item.quotation
    status = _order_tracking_status(quotation) if quotation else "pendiente"

    return {
        "item_id": item.id,
        "item_number": item_number,
        "order_label": order_label,
        "order_number": order_number,
        "quotation_id": quotation.id if quotation else None,
        "quantity": item.quantity or 0,
        "product_name": _product_name(item),
        "measure": item.measure or "—",
        "theme": item.theme or "—",
        "color": item.color or "—",
        "is_custom": _item_is_custom(item),
        "production_status": status,
        "production_status_label": TRACKING_STATUS_LABELS.get(status, status.title()),
    }


def _build_single_order_group(items: list[QuotationItem]) -> dict[str, Any]:
    first = items[0]
    quotation = first.quotation
    client = quotation.client if quotation else None
    status = _order_tracking_status(quotation) if quotation else "pendiente"
    tracking_meta = _order_tracking_meta(quotation) if quotation else {"assigned_to": "", "notes": ""}
    design_urls = get_design_urls(quotation) if quotation else []
    order_number, order_label = _order_display_ref(quotation)

    return {
        "order_number": order_number,
        "order_label": order_label,
        "production_order_id": order_number,
        "quotation_id": quotation.id if quotation else None,
        "client_name": client.name if client else "—",
        "quotation_date": quotation.created_at if quotation else None,
        "delivery_date": quotation.delivery_date if quotation else None,
        "design_urls": design_urls,
        "production_status": status,
        "production_status_label": TRACKING_STATUS_LABELS.get(status, status.title()),
        "assigned_to": tracking_meta["assigned_to"],
        "notes": tracking_meta["notes"],
        "products_count": len(items),
        "products": [
            build_condensed_row(
                item,
                order_number=order_number,
                order_label=order_label,
                item_number=index + 1,
            )
            for index, item in enumerate(items)
        ],
    }


def build_order_groups(items: list[QuotationItem]) -> list[dict[str, Any]]:
    """Agrupa ítems por cotización manteniendo el orden de aparición."""
    if not items:
        return []

    buckets: dict[int, list[QuotationItem]] = {}
    quotation_order: list[int] = []

    for item in items:
        qid = item.quotation_id
        if qid not in buckets:
            buckets[qid] = []
            quotation_order.append(qid)
        buckets[qid].append(item)

    groups: list[dict[str, Any]] = []
    for qid in quotation_order:
        bucket = sorted(buckets[qid], key=lambda row: row.id or 0)
        groups.append(_build_single_order_group(bucket))

    return groups


def paginate_order_groups(
    order_groups: list[dict[str, Any]],
    page: int,
    per_page: int,
) -> dict[str, Any]:
    import math

    page = max(1, page)
    per_page = max(1, min(per_page, 100))
    total = len(order_groups)
    pages = max(1, math.ceil(total / per_page)) if total else 1
    if page > pages:
        page = pages

    start = (page - 1) * per_page
    sliced = order_groups[start:start + per_page]

    return {
        "items": sliced,
        "page": page,
        "per_page": per_page,
        "total": total,
        "pages": pages,
        "has_prev": page > 1,
        "has_next": page < pages,
    }


def flatten_order_groups(order_groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for group in order_groups:
        for product in group["products"]:
            rows.append({
                **product,
                "order_label": group["order_label"],
                "quotation_id": group["quotation_id"],
                "client_name": group["client_name"],
                "quotation_date": group["quotation_date"],
                "delivery_date": group["delivery_date"],
                "production_status_label": group["production_status_label"],
                "production_status": group["production_status"],
            })
    return rows


def compute_kpis(db: Session, **filters) -> dict[str, int]:
    groups = get_filtered_order_groups(db, **filters)
    today = date.today()

    pending_orders = 0
    in_production_orders = 0
    ready_orders = 0
    overdue_orders = 0
    total_units = 0

    for group in groups:
        status = group["production_status"]
        delivery = group["delivery_date"]
        products_count = len(group.get("products") or [])

        if status == "pendiente":
            pending_orders += products_count
        if status in IN_PRODUCTION_STATUSES:
            in_production_orders += products_count
        if status == "listo":
            ready_orders += products_count

        if (
            delivery
            and delivery < today
            and status not in {"listo", "entregado"}
        ):
            overdue_orders += products_count

        if status != "entregado":
            total_units += sum((product["quantity"] or 0) for product in group["products"])

    return {
        "pending_products": pending_orders,
        "in_production_products": in_production_orders,
        "ready_products": ready_orders,
        "overdue_products": overdue_orders,
        "pending_units": total_units,
        "total_orders": len(groups),
        "visible_products": sum(len(group.get("products") or []) for group in groups),
    }


def update_quotation_tracking_status(
    db: Session,
    quotation_id: int,
    *,
    status: str,
    notes: str = "",
    assigned_to: str = "",
    user=None,
) -> ProductionOrder:
    from app.services.production_order_service import ensure_production_order, get_production_order_by_quotation

    quotation = db.query(Quotation).filter(Quotation.id == quotation_id).first()
    if not quotation:
        raise ValueError("Cotización no encontrada.")

    order = get_production_order_by_quotation(db, quotation_id) or ensure_production_order(db, quotation)
    if not order:
        raise ValueError("No hay orden de producción para esta cotización.")
    if not order.id:
        db.flush()

    target = normalize_status(status)
    if assigned_to:
        order.designer = assigned_to

    return transition_status(
        db,
        order,
        target,
        user=user,
        notes=notes or "",
        force=bool(user and getattr(user, "role", "") == "admin"),
    )


def get_order_detail(db: Session, quotation_id: int) -> dict[str, Any] | None:
    quotation = (
        db.query(Quotation)
        .options(
            joinedload(Quotation.client),
            joinedload(Quotation.designs),
            joinedload(Quotation.production_tracking),
            joinedload(Quotation.items).joinedload(QuotationItem.product),
        )
        .filter(Quotation.id == quotation_id)
        .first()
    )
    if not quotation:
        return None

    sync_legacy_design_file(db, quotation)
    db.commit()
    db.refresh(quotation)
    status = _order_tracking_status(quotation)
    tracking_meta = _order_tracking_meta(quotation)
    client = quotation.client

    return {
        "quotation_id": quotation.id,
        "quotation_status": quotation.status,
        "quotation_date": quotation.created_at.isoformat() if quotation.created_at else None,
        "delivery_date": quotation.delivery_date.isoformat() if quotation.delivery_date else None,
        "subtotal": quotation.subtotal,
        "total": quotation.total,
        "client_name": client.name if client else "—",
        "client_phone": client.phone if client else "",
        "client_address": client.address if client else "",
        "client_observations": client.observations if client else "",
        "design_urls": get_design_urls(quotation),
        "production_status": status,
        "production_status_label": TRACKING_STATUS_LABELS.get(status, status),
        "assigned_to": tracking_meta["assigned_to"],
        "notes": tracking_meta["notes"],
        "products": [
            {
                "item_id": item.id,
                "product_name": _product_name(item),
                "quantity": item.quantity,
                "measure": item.measure or "",
                "theme": item.theme or "",
                "color": item.color or "",
                "is_custom": _item_is_custom(item),
            }
            for item in sorted(quotation.items or [], key=lambda row: row.id or 0)
        ],
    }


def export_condensed_excel(rows: list[dict[str, Any]]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Condensado"

    headers = [
        "Orden",
        "# Producto",
        "Cotización",
        "Cliente",
        "Fecha Cotización",
        "Fecha Entrega",
        "Cantidad",
        "Producto",
        "Medida",
        "Forma",
        "Color",
        "Personalizado",
        "Estado Producción",
    ]

    header_fill = PatternFill("solid", fgColor="7C3AED")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        q_date = row["quotation_date"].strftime("%d/%m/%Y") if row["quotation_date"] else ""
        d_date = row["delivery_date"].strftime("%d/%m/%Y") if row["delivery_date"] else ""
        values = [
            row["order_label"],
            row.get("item_number", 1),
            row["quotation_id"],
            row["client_name"],
            q_date,
            d_date,
            row["quantity"],
            row["product_name"],
            row["measure"],
            row["theme"],
            row["color"],
            "Sí" if row["is_custom"] else "No",
            row["production_status_label"],
        ]
        for col, value in enumerate(values, start=1):
            sheet.cell(row=row_idx, column=col, value=value)

    for col in sheet.columns:
        sheet.column_dimensions[col[0].column_letter].width = 16

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def export_condensed_pdf(rows: list[dict[str, Any]], **kwargs) -> BytesIO:
    from app.services.condensed_pdf import export_condensed_production_pdf

    return export_condensed_production_pdf(rows, **kwargs)
