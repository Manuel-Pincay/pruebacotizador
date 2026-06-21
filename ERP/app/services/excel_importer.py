from openpyxl import load_workbook
from openpyxl import Workbook

from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

from openpyxl.utils import get_column_letter

from io import BytesIO
from datetime import datetime

from sqlalchemy.orm import Session

from app.models.client import Client
from app.models.product import Product
from app.models.quotation import Quotation
from app.models.quotation_item import QuotationItem
from app.services.logo_types import logo_type_label, resolve_item_logo_type
from app.utils.text_format import format_title_words


# ==========================================
# STYLES
# ==========================================

HEADER_FILL = PatternFill(
    start_color="7C3AED",
    end_color="7C3AED",
    fill_type="solid"
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True
)

DESCRIPTION_FILL = PatternFill(
    start_color="E9D5FF",
    end_color="E9D5FF",
    fill_type="solid"
)

CENTER = Alignment(
    horizontal="center",
    vertical="center"
)


# ==========================================
# FORMAT SHEET
# ==========================================

def format_sheet(ws):

    # HEADER STYLE

    for cell in ws[1]:

        cell.fill = HEADER_FILL

        cell.font = HEADER_FONT

        cell.alignment = CENTER

    # DESCRIPTION STYLE

    for cell in ws[2]:

        cell.fill = DESCRIPTION_FILL

        cell.alignment = CENTER

    # AUTO WIDTH

    for column_cells in ws.columns:

        length = max(

            len(str(cell.value or ""))

            for cell in column_cells

        )

        col_letter = get_column_letter(
            column_cells[0].column
        )

        ws.column_dimensions[
            col_letter
        ].width = length + 8

    # FREEZE HEADER

    ws.freeze_panes = "A2"


CLIENT_HEADERS = [
    "name",
    "company",
    "ruc_ci",
    "phone",
    "email",
    "address",
    "client_type",
    "observations",
]

CLIENT_DESCRIPTIONS = [
    "Nombre cliente",
    "Empresa",
    "RUC o CI",
    "Teléfono",
    "Correo",
    "Dirección",
    "Mayorista / Minorista",
    "Notas adicionales",
]

PRODUCT_HEADERS = [
    "code",
    "name",
    "description",
    "category",
    "material",
    "color",
    "size",
    "thickness",
    "price",
    "cost",
    "theme",
    "stock",
    "custom",
]

PRODUCT_DESCRIPTIONS = [
    "Código producto",
    "Nombre producto",
    "Descripción",
    "Categoría",
    "Material",
    "Color",
    "Tamaño",
    "Espesor",
    "Precio venta",
    "Costo",
    "Temática",
    "Stock",
    "TRUE/FALSE",
]


def _workbook_to_buffer(wb: Workbook) -> BytesIO:
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_clients_excel(db: Session) -> BytesIO:
    """Exporta clientes en formato compatible con la plantilla de importación."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"

    ws.append(CLIENT_HEADERS)
    ws.append(CLIENT_DESCRIPTIONS)

    clients = db.query(Client).order_by(Client.name.asc()).all()
    for client in clients:
        ws.append([
            client.name or "",
            client.company or "",
            client.ruc_ci or "",
            client.phone or "",
            client.email or "",
            client.address or "",
            client.client_type or "",
            client.observations or "",
        ])

    format_sheet(ws)
    return _workbook_to_buffer(wb)


def export_products_excel(db: Session) -> BytesIO:
    """Exporta productos en formato compatible con la plantilla de importación."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    ws.append(PRODUCT_HEADERS)
    ws.append(PRODUCT_DESCRIPTIONS)

    products = db.query(Product).order_by(Product.name.asc()).all()
    for product in products:
        ws.append([
            product.code or "",
            product.name or "",
            product.description or "",
            product.category or "",
            product.material or "",
            product.color or "",
            product.size or "",
            product.thickness or "",
            product.price if product.price is not None else 0,
            product.cost if product.cost is not None else 0,
            product.theme or "",
            product.stock if product.stock is not None else 0,
            "TRUE" if product.custom else "FALSE",
        ])

    format_sheet(ws)
    return _workbook_to_buffer(wb)


QUOTATION_SUMMARY_HEADERS = [
    "quotation_id",
    "created_at",
    "delivery_date",
    "status",
    "client_name",
    "client_ruc_ci",
    "client_phone",
    "subtotal",
    "discount_percent",
    "discount_amount",
    "iva_percent",
    "iva_amount",
    "shipping_cost",
    "total",
    "total_paid",
    "pending_balance",
    "payment_status",
]

QUOTATION_SUMMARY_DESCRIPTIONS = [
    "N° Cotización",
    "Fecha emisión",
    "Fecha entrega",
    "Estado",
    "Cliente",
    "CI/RUC",
    "Teléfono",
    "Subtotal",
    "Descuento %",
    "Descuento $",
    "IVA %",
    "IVA $",
    "Envío",
    "Total",
    "Abonado",
    "Saldo",
    "Estado pago",
]

QUOTATION_ITEM_HEADERS = [
    "quotation_id",
    "client_name",
    "quotation_date",
    "quotation_status",
    "item_id",
    "detail",
    "quantity",
    "unit_price",
    "item_discount_percent",
    "line_total",
    "measure",
    "theme",
    "color",
    "logo_type",
]

QUOTATION_ITEM_DESCRIPTIONS = [
    "N° Cotización",
    "Cliente",
    "Fecha cotización",
    "Estado cotización",
    "N° Ítem",
    "Detalle",
    "Cantidad",
    "Precio unit.",
    "Descuento ítem %",
    "Total línea",
    "Medida",
    "Temática",
    "Color",
    "Tipo logo",
]

PAYMENT_STATUS_LABELS = {
    "sin_abono": "Sin abono",
    "parcial": "Parcial",
    "pagada": "Pagada",
}


def _fmt_date(value) -> str:
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y")
    return str(value)


def _fmt_datetime(value) -> str:
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y %H:%M")
    return str(value)


def export_quotations_excel(db: Session) -> BytesIO:
    """Exporta cotizaciones (resumen + detalle de ítems) a Excel."""
    from sqlalchemy.orm import joinedload

    wb = Workbook()

    # ── Hoja 1: Resumen cotizaciones ──
    ws_summary = wb.active
    ws_summary.title = "Cotizaciones"
    ws_summary.append(QUOTATION_SUMMARY_HEADERS)
    ws_summary.append(QUOTATION_SUMMARY_DESCRIPTIONS)

    # ── Hoja 2: Detalle ítems ──
    ws_items = wb.create_sheet("Detalle")
    ws_items.append(QUOTATION_ITEM_HEADERS)
    ws_items.append(QUOTATION_ITEM_DESCRIPTIONS)

    quotations = (
        db.query(Quotation)
        .options(
            joinedload(Quotation.client),
            joinedload(Quotation.items).joinedload(QuotationItem.product),
            joinedload(Quotation.payments),
        )
        .order_by(Quotation.created_at.desc(), Quotation.id.desc())
        .all()
    )

    for quotation in quotations:
        client = quotation.client
        subtotal = float(quotation.subtotal or 0)
        discount_pct = float(quotation.discount or 0)
        discount_amt = subtotal * (discount_pct / 100)
        subtotal_after = subtotal - discount_amt
        iva_pct = float(quotation.iva or 0)
        iva_amt = subtotal_after * (iva_pct / 100)
        shipping = float(quotation.shipping_cost or 0)
        total_paid = quotation.total_paid
        pending = quotation.pending_balance
        pay_status = PAYMENT_STATUS_LABELS.get(
            quotation.payment_status,
            quotation.payment_status or "",
        )

        ws_summary.append([
            quotation.id,
            _fmt_datetime(quotation.created_at),
            _fmt_date(quotation.delivery_date),
            quotation.status or "",
            client.name if client else "",
            client.ruc_ci if client else "",
            client.phone if client else "",
            round(subtotal, 2),
            round(discount_pct, 2),
            round(discount_amt, 2),
            round(iva_pct, 2),
            round(iva_amt, 2),
            round(shipping, 2),
            round(float(quotation.total or 0), 2),
            round(total_paid, 2),
            round(pending, 2),
            pay_status,
        ])

        client_name = client.name if client else ""
        q_date = _fmt_datetime(quotation.created_at)
        q_status = quotation.status or ""

        for item in quotation.items:
            item_discount = float(getattr(item, "item_discount", 0) or 0)
            ws_items.append([
                quotation.id,
                client_name,
                q_date,
                q_status,
                item.id,
                item.detail or (item.product.name if item.product else ""),
                item.quantity or 0,
                round(float(item.unit_price or 0), 2),
                round(item_discount, 2),
                round(float(item.total or 0), 2),
                item.measure or "",
                item.theme or "",
                item.color or "",
                logo_type_label(resolve_item_logo_type(item)),
            ])

    format_sheet(ws_summary)
    format_sheet(ws_items)
    return _workbook_to_buffer(wb)


def export_filename(prefix: str) -> str:
    stamp = datetime.now().strftime("%Y-%m-%d")
    return f"{prefix}_{stamp}.xlsx"


# ==========================================
# CLIENT TEMPLATE
# ==========================================

def create_clients_template(path):

    wb = Workbook()

    ws = wb.active

    ws.title = "Clientes"

    # HEADERS

    headers = CLIENT_HEADERS

    ws.append(headers)

    # DESCRIPTIONS

    ws.append(CLIENT_DESCRIPTIONS)

    # EXAMPLE

    ws.append([

        "Juan Pérez",
        "Innova Arte",
        "1723456789",
        "0999999999",
        "juan@gmail.com",
        "Quito",
        "Mayorista",
        "Cliente frecuente"

    ])

    format_sheet(ws)

    wb.save(path)


# ==========================================
# PRODUCT TEMPLATE
# ==========================================

def create_products_template(path):

    wb = Workbook()

    ws = wb.active

    ws.title = "Productos"

    # HEADERS

    headers = PRODUCT_HEADERS

    ws.append(headers)

    # DESCRIPTIONS

    ws.append(PRODUCT_DESCRIPTIONS)

    # EXAMPLE

    ws.append([

        "TOP001",
        "Topper Feliz Cumpleaños",
        "Topper MDF dorado",
        "Topper",
        "MDF",
        "Dorado",
        "30cm",
        "3mm",
        "5.50",
        "2.00",
        "Feliz Cumpleaños",
        "50",
        "TRUE"

    ])

    format_sheet(ws)

    wb.save(path)


# ==========================================
# IMPORT CLIENTS
# ==========================================

def import_clients(db, file_path):

    wb = load_workbook(file_path)

    ws = wb.active

    imported = 0

    for row in ws.iter_rows(
        min_row=3,
        values_only=True
    ):

        if not row[0]:
            continue

        client = Client(

            name=format_title_words(str(row[0] or "")),
            company=format_title_words(str(row[1] or "")) or None,
            ruc_ci=row[2],
            phone=row[3],
            email=row[4],
            address=format_title_words(str(row[5] or "")) or None,
            client_type=row[6],
            observations=row[7]

        )

        db.add(client)

        imported += 1

    db.commit()

    return imported


# ==========================================
# IMPORT PRODUCTS
# ==========================================

def import_products(db, file_path):

    wb = load_workbook(file_path)

    ws = wb.active

    imported = 0

    for row in ws.iter_rows(
        min_row=3,
        values_only=True
    ):

        if not row[1]:
            continue

        product = Product(

            code=row[0],
            name=format_title_words(str(row[1] or "")),
            description=row[2],
            category=format_title_words(str(row[3] or "")),
            material=format_title_words(str(row[4] or "")),
            color=format_title_words(str(row[5] or "")),
            size=row[6],
            thickness=row[7],
            price=float(row[8] or 0),
            cost=float(row[9] or 0),
            theme=format_title_words(str(row[10] or "")),
            stock=int(row[11] or 0),
            custom=str(row[12]).lower() == "true"

        )

        db.add(product)

        imported += 1

    db.commit()

    return imported