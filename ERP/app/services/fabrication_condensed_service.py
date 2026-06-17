"""Condensado de fabricación: órdenes con datos de impresión para todo el equipo."""

from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app.models.client import Client
from app.models.product import Product
from app.models.production_order import ProductionOrder
from app.models.quotation import Quotation
from app.models.quotation_item import QuotationItem
from app.models.user import User
from app.services.condensed_service import (
    TRACKING_STATUS_LABELS,
    annotate_group_sections,
    build_condensed_row,
    build_order_groups,
    filter_order_groups,
    paginate_order_groups,
    sort_order_groups,
)
from app.services.production_helpers import order_delivery_date, production_orders_base_query
from app.services.production_order_service import (
    PRODUCTION_STATUS_LABELS,
    can_transition,
    get_production_order,
    normalize_status,
    transition_status,
)
from app.services.quotation_design_service import get_design_urls, sync_legacy_design_file

COMPLETED_FABRICATION_STATUSES = {"envio", "entregado", "cancelado"}


def _load_fabrication_orders(db: Session) -> list[ProductionOrder]:
    return (
        production_orders_base_query(db)
        .options(
            joinedload(ProductionOrder.quotation).joinedload(Quotation.client),
        )
        .order_by(ProductionOrder.id.asc())
        .all()
    )


def _fabrication_meta(order: ProductionOrder) -> dict[str, Any]:
    return {
        "production_order_id": order.id,
        "order_label": f"OP-{order.id:04d}",
        "file_name": order.design_file_name or "",
        "material": order.design_material or "",
        "size": order.design_size or "",
        "usb_reference": order.design_usb_reference or "",
        "copies": order.design_copies or 0,
        "notes": order.design_notes or "",
        "fabricator": order.fabricator or "",
        "designer": order.designer or "",
        "print_url": f"/production/{order.id}/fabrication/print",
        "detail_url": f"/production/{order.id}",
    }


def parse_fabrication_filters(
    *,
    month: str = "",
    year: str = "",
    client_id: str = "",
    production_status: str = "",
    custom_filter: str = "",
    search: str = "",
    group_by: str = "order",
    show_completed: str = "",
    apply_current_month_default: bool = False,
) -> dict[str, Any]:
    today = date.today()
    parsed_month = int(month) if month.strip().isdigit() else None
    parsed_year = int(year) if year.strip().isdigit() else None

    if apply_current_month_default and parsed_month is None and parsed_year is None:
        parsed_month = today.month
        parsed_year = today.year
    elif parsed_month and not parsed_year:
        parsed_year = today.year

    parsed_client = int(client_id) if client_id.strip().isdigit() else None

    return {
        "month": parsed_month,
        "year": parsed_year,
        "client_id": parsed_client,
        "production_status": production_status.strip(),
        "custom_filter": custom_filter.strip(),
        "search": search.strip(),
        "group_by": group_by if group_by in {"order", "delivery", "client", "product"} else "order",
        "show_completed": show_completed.strip().lower() in {"1", "true", "yes", "on"},
    }


def _order_in_month(order: ProductionOrder, month: int | None, year: int | None) -> bool:
    if not month and not year:
        return True
    delivery = order_delivery_date(order)
    if not delivery:
        return False
    target_year = year or date.today().year
    if year and delivery.year != target_year:
        return False
    if month and delivery.month != month:
        return False
    return True


def _build_items_query(db: Session, quotation_ids: set[int]):
    return (
        db.query(QuotationItem)
        .join(Quotation, QuotationItem.quotation_id == Quotation.id)
        .join(Client, Quotation.client_id == Client.id)
        .outerjoin(Product, QuotationItem.product_id == Product.id)
        .filter(QuotationItem.quotation_id.in_(quotation_ids))
        .options(
            joinedload(QuotationItem.quotation).joinedload(Quotation.client),
            joinedload(QuotationItem.quotation).joinedload(Quotation.designs),
            joinedload(QuotationItem.quotation).joinedload(Quotation.production_order),
            joinedload(QuotationItem.product),
        )
        .order_by(Quotation.delivery_date.desc(), Quotation.id.desc(), QuotationItem.id.desc())
    )


def _fabricator_action_flags(status: str) -> dict[str, bool]:
    st = normalize_status(status)
    return {
        "can_mark_taken": st == "diseno" and can_transition(st, "produccion"),
        "can_mark_completed": st == "produccion" and can_transition(st, "envio"),
    }


def _enrich_groups(
    groups: list[dict[str, Any]],
    orders_by_quotation: dict[int, ProductionOrder],
) -> list[dict[str, Any]]:
    enriched: list[dict[str, Any]] = []
    for group in groups:
        order = orders_by_quotation.get(group.get("quotation_id"))
        if not order:
            continue
        fabrication = _fabrication_meta(order)
        delivery = order_delivery_date(order)
        updated_products = []
        for product in group.get("products") or []:
            row = dict(product)
            row["order_label"] = fabrication["order_label"]
            row["order_number"] = order.id
            updated_products.append(row)

        enriched.append({
            **group,
            **fabrication,
            **_fabricator_action_flags(order.status),
            "delivery_date": delivery,
            "products": updated_products,
            "products_count": len(updated_products),
        })
    return enriched


def get_fabrication_order_groups(
    db: Session,
    *,
    user: User | None = None,
    admin_view: bool = False,
    **filters,
) -> list[dict[str, Any]]:
    orders = _load_fabrication_orders(db)

    if filters.get("month") or filters.get("year"):
        orders = [
            order for order in orders
            if _order_in_month(order, filters.get("month"), filters.get("year"))
        ]

    if filters.get("client_id"):
        orders = [
            order for order in orders
            if order.quotation and order.quotation.client_id == filters["client_id"]
        ]

    if not filters.get("show_completed"):
        orders = [
            order for order in orders
            if normalize_status(order.status) not in COMPLETED_FABRICATION_STATUSES
        ]

    quotation_ids = {order.quotation_id for order in orders if order.quotation_id}
    if not quotation_ids:
        return []

    items = _build_items_query(db, quotation_ids).all()
    groups = build_order_groups(items)
    orders_by_quotation = {order.quotation_id: order for order in orders}
    groups = _enrich_groups(groups, orders_by_quotation)

    groups = filter_order_groups(
        groups,
        production_status=filters.get("production_status", ""),
        custom_filter=filters.get("custom_filter", ""),
        search=filters.get("search", ""),
    )
    groups = sort_order_groups(groups, filters.get("group_by", "order"))
    return annotate_group_sections(groups, filters.get("group_by", "order"))


def compute_fabrication_kpis(groups: list[dict[str, Any]]) -> dict[str, int]:
    today = date.today()
    pending_products = 0
    in_production_products = 0
    ready_products = 0
    overdue_products = 0
    pending_units = 0

    for group in groups:
        status = (group.get("production_status") or "pendiente").lower()
        delivery = group.get("delivery_date")
        products = group.get("products") or []
        count = len(products)

        if status == "pendiente":
            pending_products += count
        if status in {"diseno", "produccion"}:
            in_production_products += count
        if status in {"envio", "entregado"}:
            ready_products += count
        if delivery and delivery < today and status not in {"entregado", "cancelado"}:
            overdue_products += count
        if status != "entregado":
            pending_units += sum(int(product.get("quantity") or 0) for product in products)

    return {
        "pending_products": pending_products,
        "in_production_products": in_production_products,
        "ready_products": ready_products,
        "overdue_products": overdue_products,
        "pending_units": pending_units,
        "total_orders": len(groups),
        "visible_products": sum(len(group.get("products") or []) for group in groups),
    }


def active_filter_labels(filters: dict[str, Any], clients: list) -> list[str]:
    labels: list[str] = []
    month_names = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
    }

    if filters.get("month") or filters.get("year"):
        if filters.get("month") and filters.get("year"):
            labels.append(
                f"{month_names.get(filters['month'], filters['month'])} {filters['year']} (entrega)"
            )
        elif filters.get("year"):
            labels.append(f"Año {filters['year']} (entrega)")
        elif filters.get("month"):
            labels.append(f"{month_names.get(filters['month'], filters['month'])} (entrega)")

    if filters.get("client_id"):
        client_name = next(
            (client.name for client in clients if client.id == filters["client_id"]),
            f"Cliente #{filters['client_id']}",
        )
        labels.append(f"Cliente: {client_name}")

    if filters.get("production_status"):
        labels.append(
            "Estado: "
            + TRACKING_STATUS_LABELS.get(
                filters["production_status"],
                filters["production_status"],
            )
        )

    if filters.get("custom_filter") == "yes":
        labels.append("Con personalizados")
    elif filters.get("custom_filter") == "no":
        labels.append("Solo catálogo")

    if filters.get("search"):
        labels.append(f'Búsqueda: "{filters["search"]}"')

    if filters.get("show_completed"):
        labels.append("Incluye ya realizadas (envío/entregadas)")
    else:
        labels.append("Solo pendientes de fabricar")

    return labels


def list_fabricator_names(db: Session) -> list[str]:
    rows = (
        db.query(ProductionOrder.fabricator)
        .filter(ProductionOrder.fabricator.isnot(None))
        .distinct()
        .all()
    )
    names = sorted(
        {(row[0] or "").strip() for row in rows if (row[0] or "").strip()},
        key=str.lower,
    )
    return names


PENDING_FABRICATION_DASHBOARD_STATUSES = {"pendiente", "diseno", "produccion"}
DONE_FABRICATION_DASHBOARD_STATUSES = {"envio", "entregado"}


def build_fabricator_dashboard(db: Session, user: User | None = None) -> dict[str, Any]:
    """Resumen simple para el dashboard del rol producción."""
    orders = _load_fabrication_orders(db)
    pending: list[dict[str, Any]] = []
    done: list[dict[str, Any]] = []

    for order in orders:
        st = normalize_status(order.status)
        if st == "cancelado":
            continue
        client = order.quotation.client if order.quotation else None
        delivery = order_delivery_date(order)
        row = {
            "id": order.id,
            "label": f"OP-{order.id:04d}",
            "client_name": client.name if client else "—",
            "status": st,
            "status_label": PRODUCTION_STATUS_LABELS.get(st, st),
            "file_name": order.design_file_name or "—",
            "delivery_label": delivery.strftime("%d/%m/%Y") if delivery else "—",
        }
        if st in PENDING_FABRICATION_DASHBOARD_STATUSES:
            pending.append(row)
        elif st in DONE_FABRICATION_DASHBOARD_STATUSES:
            done.append(row)

    pending.sort(key=lambda item: item["delivery_label"])
    done.sort(key=lambda item: item["id"], reverse=True)

    return {
        "pending_count": len(pending),
        "done_count": len(done),
        "pending_orders": pending[:10],
        "done_orders": done[:10],
    }


def assert_fabricator_can_manage_order(
    order: ProductionOrder,
    user: User,
) -> None:
    if getattr(user, "role", "") not in {"admin", "produccion"}:
        raise ValueError("No tienes permiso para gestionar esta orden.")


def fabricator_mark_taken(db: Session, order_id: int, user: User) -> ProductionOrder:
    order = get_production_order(db, order_id)
    if not order:
        raise ValueError("Orden no encontrada.")
    assert_fabricator_can_manage_order(order, user)

    current = normalize_status(order.status)
    if current == "produccion":
        return order
    if current == "diseno":
        return transition_status(
            db,
            order,
            "produccion",
            user=user,
            notes="Tomada por fabricador.",
        )
    raise ValueError(
        f"No se puede marcar como tomada desde estado "
        f"{PRODUCTION_STATUS_LABELS.get(current, current)}."
    )


def fabricator_mark_completed(db: Session, order_id: int, user: User) -> ProductionOrder:
    order = get_production_order(db, order_id)
    if not order:
        raise ValueError("Orden no encontrada.")
    assert_fabricator_can_manage_order(order, user)

    current = normalize_status(order.status)
    if current != "produccion":
        raise ValueError(
            "Solo órdenes en producción pueden marcarse como realizadas."
        )
    return transition_status(
        db,
        order,
        "envio",
        user=user,
        notes="Fabricación completada — pasa a envío.",
    )


def get_fabrication_order_detail(db: Session, quotation_id: int) -> dict[str, Any] | None:
    quotation = (
        db.query(Quotation)
        .options(
            joinedload(Quotation.client),
            joinedload(Quotation.designs),
            joinedload(Quotation.production_order),
            joinedload(Quotation.items).joinedload(QuotationItem.product),
        )
        .filter(Quotation.id == quotation_id)
        .first()
    )
    if not quotation or not quotation.production_order:
        return None

    sync_legacy_design_file(db, quotation)
    db.commit()
    db.refresh(quotation)

    order = quotation.production_order
    client = quotation.client
    status = normalize_status(order.status)

    return {
        "quotation_id": quotation.id,
        "production_order_id": order.id,
        "order_label": f"OP-{order.id:04d}",
        "quotation_date": quotation.created_at.isoformat() if quotation.created_at else None,
        "delivery_date": quotation.delivery_date.isoformat() if quotation.delivery_date else None,
        "client_name": client.name if client else "—",
        "design_urls": get_design_urls(quotation),
        "production_status": status,
        "production_status_label": TRACKING_STATUS_LABELS.get(status, status),
        "fabrication": _fabrication_meta(order),
        "products": [
            build_condensed_row(
                item,
                order_number=order.id,
                order_label=f"OP-{order.id:04d}",
                item_number=index + 1,
            )
            for index, item in enumerate(sorted(quotation.items or [], key=lambda row: row.id or 0))
        ],
    }


def flatten_fabrication_groups(groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for group in groups:
        fabrication = {
            "file_name": group.get("file_name") or "",
            "material": group.get("material") or "",
            "size": group.get("size") or "",
            "usb_reference": group.get("usb_reference") or "",
            "copies": group.get("copies") or 0,
            "fabricator": group.get("fabricator") or "",
            "designer": group.get("designer") or "",
        }
        for product in group.get("products") or []:
            rows.append({
                **product,
                **fabrication,
                "order_label": group.get("order_label"),
                "quotation_id": group.get("quotation_id"),
                "client_name": group.get("client_name"),
                "delivery_date": group.get("delivery_date"),
                "production_status_label": group.get("production_status_label"),
                "production_status": group.get("production_status"),
            })
    return rows


def export_fabrication_excel(groups: list[dict[str, Any]]) -> BytesIO:
    """Una fila por orden; productos como resumen compacto."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Fabricación"

    headers = [
        "Orden", "Cotización", "Cliente", "Entrega", "Archivo", "Material",
        "Medida imp.", "USB", "Copias", "Diseñador", "Productos (ref.)", "Estado",
    ]
    header_fill = PatternFill("solid", fgColor="EA580C")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, group in enumerate(groups, start=2):
        delivery = group.get("delivery_date")
        delivery_label = delivery.strftime("%d/%m/%Y") if isinstance(delivery, date) else "—"
        products = group.get("products") or []
        products_summary = "; ".join(
            f"{p.get('quantity', 0)}× {p.get('product_name', '—')}"
            for p in products
        ) or "—"
        values = [
            group.get("order_label") or "",
            group.get("quotation_id") or "",
            group.get("client_name") or "",
            delivery_label,
            group.get("file_name") or "",
            group.get("material") or "",
            group.get("size") or "",
            group.get("usb_reference") or "",
            group.get("copies") or 0,
            group.get("designer") or "",
            products_summary,
            group.get("production_status_label") or "",
        ]
        for col, value in enumerate(values, start=1):
            sheet.cell(row=row_idx, column=col, value=value)

    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 14

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def export_fabrication_pdf(groups: list[dict[str, Any]], **kwargs) -> BytesIO:
    from app.services.fabrication_condensed_pdf import build_fabrication_condensed_pdf

    return build_fabrication_condensed_pdf(order_groups=groups, **kwargs)
